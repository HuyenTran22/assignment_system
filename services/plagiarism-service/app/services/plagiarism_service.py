"""
Advanced plagiarism detection service using text similarity.

Supports multiple file formats: TXT, PDF, DOCX, XLSX
Compares all files in submissions, not just the first file.

For production, consider integrating with:
- Turnitin API
- Copyscape API
- Or more sophisticated algorithms (LSH, MinHash, etc.)
"""

from pathlib import Path
from typing import List, Tuple, Optional
from decimal import Decimal
import re
import os


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file."""
    try:
        try:
            import PyPDF2
        except ImportError:
            print(f"PyPDF2 not installed. Install with: pip install PyPDF2")
            return ""
        text = ""
        with open(file_path, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Error reading PDF {file_path}: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX file."""
    try:
        try:
            from docx import Document
        except ImportError:
            print(f"python-docx not installed. Install with: pip install python-docx")
            return ""
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    except Exception as e:
        print(f"Error reading DOCX {file_path}: {e}")
        return ""


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from XLSX file."""
    try:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print(f"openpyxl not installed. Install with: pip install openpyxl")
            return ""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        print(f"Error reading XLSX {file_path}: {e}")
        return ""


def extract_text_from_file(file_path: str) -> str:
    """
    Extract text from a file based on its extension.
    
    Supports:
    - TXT: Plain text files
    - PDF: PDF documents (using PyPDF2)
    - DOCX: Word documents (using python-docx)
    - XLSX: Excel files (using openpyxl)
    """
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return ""
    
    file_ext = Path(file_path).suffix.lower()
    
    try:
        if file_ext == '.pdf':
            return extract_text_from_pdf(file_path)
        elif file_ext in ['.docx', '.doc']:
            return extract_text_from_docx(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return extract_text_from_xlsx(file_path)
        elif file_ext == '.txt' or file_ext == '':
            # Try to read as text file
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        else:
            # Try to read as text file for unknown extensions
            print(f"Unknown file extension {file_ext}, attempting to read as text")
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return ""


def preprocess_text(text: str) -> str:
    """Normalize text for comparison."""
    # Convert to lowercase
    text = text.lower()
    
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    
    # Remove special characters (keep alphanumeric and spaces)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    
    return text.strip()


def tokenize(text: str) -> List[str]:
    """Split text into words."""
    return text.split()


def calculate_jaccard_similarity(text1: str, text2: str) -> float:
    """
    Calculate Jaccard similarity between two texts.
    
    Jaccard similarity = |A ∩ B| / |A ∪ B|
    
    Returns: Similarity score (0-100)
    """
    # Preprocess
    text1 = preprocess_text(text1)
    text2 = preprocess_text(text2)
    
    # Tokenize
    tokens1 = set(tokenize(text1))
    tokens2 = set(tokenize(text2))
    
    # Handle empty texts
    if not tokens1 or not tokens2:
        return 0.0
    
    # Calculate Jaccard similarity
    intersection = tokens1.intersection(tokens2)
    union = tokens1.union(tokens2)
    
    if not union:
        return 0.0
    
    similarity = len(intersection) / len(union)
    
    # Convert to percentage
    return round(similarity * 100, 2)


def calculate_cosine_similarity(text1: str, text2: str) -> float:
    """
    Calculate cosine similarity between two texts.
    
    More sophisticated than Jaccard, considers word frequency.
    
    Returns: Similarity score (0-100)
    """
    from collections import Counter
    import math
    
    # Preprocess
    text1 = preprocess_text(text1)
    text2 = preprocess_text(text2)
    
    # Tokenize and count
    tokens1 = tokenize(text1)
    tokens2 = tokenize(text2)
    
    if not tokens1 or not tokens2:
        return 0.0
    
    # Create frequency vectors
    counter1 = Counter(tokens1)
    counter2 = Counter(tokens2)
    
    # Get all unique words
    all_words = set(counter1.keys()).union(set(counter2.keys()))
    
    # Create vectors
    vec1 = [counter1.get(word, 0) for word in all_words]
    vec2 = [counter2.get(word, 0) for word in all_words]
    
    # Calculate dot product
    dot_product = sum(a * b for a, b in zip(vec1, vec2))
    
    # Calculate magnitudes
    magnitude1 = math.sqrt(sum(a * a for a in vec1))
    magnitude2 = math.sqrt(sum(b * b for b in vec2))
    
    if magnitude1 == 0 or magnitude2 == 0:
        return 0.0
    
    # Calculate cosine similarity
    similarity = dot_product / (magnitude1 * magnitude2)
    
    # Convert to percentage
    return round(similarity * 100, 2)


def compare_submissions(file_path1: str, file_path2: str) -> Decimal:
    """
    Compare two submission files and return similarity score.
    
    Uses cosine similarity for better accuracy.
    
    Returns: Similarity score (0-100) as Decimal
    """
    # Extract text
    text1 = extract_text_from_file(file_path1)
    text2 = extract_text_from_file(file_path2)
    
    # If either file is empty or too short, return 0
    if len(text1) < 50 or len(text2) < 50:
        return Decimal("0.00")
    
    # Calculate similarity
    similarity = calculate_cosine_similarity(text1, text2)
    
    return Decimal(str(similarity))


def compare_all_files_in_submissions(
    submission1_files: List[Tuple[str, str]], 
    submission2_files: List[Tuple[str, str]]
) -> Decimal:
    """
    Compare all files between two submissions and return maximum similarity.
    
    Args:
        submission1_files: List of (file_id, file_path) tuples for submission 1
        submission2_files: List of (file_id, file_path) tuples for submission 2
    
    Returns:
        Maximum similarity score (0-100) as Decimal
    """
    max_similarity = Decimal("0.00")
    
    # Compare each file from submission1 with each file from submission2
    for file1_id, path1 in submission1_files:
        for file2_id, path2 in submission2_files:
            similarity = compare_submissions(path1, path2)
            if similarity > max_similarity:
                max_similarity = similarity
    
    return max_similarity


def compare_all_submissions(
    submission_data: List[Tuple[str, str, List[Tuple[str, str]]]]
) -> List[Tuple[str, str, Decimal]]:
    """
    Compare all pairs of submissions, comparing all files in each submission.
    
    Args:
        submission_data: List of (submission_id, student_id, files) tuples
            where files is List of (file_id, file_path) tuples
    
    Returns:
        List of (submission1_id, submission2_id, similarity_score) tuples
        where similarity_score is the maximum similarity found between any files
    """
    results = []
    
    # Compare each pair of submissions
    for i in range(len(submission_data)):
        for j in range(i + 1, len(submission_data)):
            sub1_id, student1_id, files1 = submission_data[i]
            sub2_id, student2_id, files2 = submission_data[j]
            
            # Don't compare submissions from the same student
            if student1_id == student2_id:
                continue
            
            # Skip if either submission has no files
            if not files1 or not files2:
                continue
            
            # Compare all files between the two submissions
            similarity = compare_all_files_in_submissions(files1, files2)
            
            # Only include if similarity > 0
            if similarity > 0:
                results.append((sub1_id, sub2_id, similarity))
    
    return results
